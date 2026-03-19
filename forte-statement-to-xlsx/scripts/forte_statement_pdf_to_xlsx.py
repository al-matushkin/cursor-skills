#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Iterable, Optional


def _parse_number(s: str) -> float:
    # Handles "1 252 527.52" and "2012527.52" and "1,23" formats.
    s = s.strip().replace(" ", "")
    s = s.replace(",", ".")
    return float(s)


def extract_final_balance_kzt(pdf_path: Path) -> Optional[float]:
    # We only need one value; scan extracted text lines for the "Доступно на ... KZT" pattern.
    import pdfplumber  # type: ignore

    final_re = re.compile(
        r"Доступно на\s*[^:]+:\s*([0-9][0-9\s\.,]*)\s*KZT", re.IGNORECASE
    )
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                m = final_re.search(line)
                if m:
                    return _parse_number(m.group(1))
    return None


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Convert a Forte/ForteForex KZT statement PDF to XLSX with cancelled-order-aware net."
    )
    ap.add_argument("pdf", type=Path, help="Input PDF path")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output XLSX path (default: next to PDF with .xlsx extension)",
    )
    ap.add_argument(
        "--initial-balance-kzt",
        type=float,
        default=None,
        help="Initial KZT balance at period start (optional; used to compute expected final).",
    )
    args = ap.parse_args()

    pdf_path: Path = args.pdf
    out_path: Path = args.output or pdf_path.with_suffix(".xlsx")

    # Import local extraction helpers from this skill folder scripts/
    scripts_dir = Path(__file__).resolve().parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    from pdf_to_xlsx import (  # type: ignore
        fix_misattached_merchants,
        iter_pdf_lines,
        parse_transactions,
        to_dataframe,
    )

    # 1) Extract transactions into DataFrame
    lines: Iterable[str] = iter_pdf_lines(pdf_path)
    txns = parse_transactions(lines)
    txns = fix_misattached_merchants(txns)
    df = to_dataframe(txns)

    required_cols = {"date", "amount", "currency", "description", "details"}
    missing = required_cols.difference(df.columns)
    if missing:
        raise SystemExit(f"Missing expected columns from extractor: {sorted(missing)}")

    # 2) Detect cancelled ForteForex bundles (same-day match by values)
    df["is_cancelled_forteforex_bundle"] = 0
    df["date_only"] = df["date"].dt.date
    df["_text"] = (
        df["description"].fillna("").astype(str) + " " + df["details"].fillna("").astype(str)
    ).str.strip()

    principal_kw = "Снятие заявки по ForteForex"
    commission_kw = "Отмена комиссии по сделке ForteForex"

    is_principal = (
        df["_text"].str.contains(principal_kw, case=False, regex=False) & (df["amount"] > 0)
    )
    is_commission = (
        df["_text"].str.contains(commission_kw, case=False, regex=False) & (df["amount"] > 0)
    )
    # Initial ForteForex debit leg: negative KZT with "Списание".
    is_debit = df["_text"].str.contains("Списание", case=False, regex=False) & (df["amount"] < 0)

    for day in df["date_only"].dropna().unique():
        day_df = df[df["date_only"] == day]
        if day_df.empty:
            continue

        principal_idx = day_df[is_principal.loc[day_df.index]].index.tolist()
        commission_idx = day_df[is_commission.loc[day_df.index]].index.tolist()
        debit_idx = day_df[is_debit.loc[day_df.index]].index.tolist()

        if not principal_idx or not commission_idx or not debit_idx:
            continue

        used_pr: set = set()
        used_com: set = set()
        used_deb: set = set()

        deb_by_amt: dict[float, list] = {}
        for di in debit_idx:
            amt = float(df.at[di, "amount"])
            deb_by_amt.setdefault(round(amt, 2), []).append(di)

        def find_debit(expected_initial: float):
            r2 = round(expected_initial, 2)
            for di in deb_by_amt.get(r2, []):
                if di not in used_deb:
                    return di
            return None

        for pi in principal_idx:
            if pi in used_pr:
                continue
            p_amt = float(df.at[pi, "amount"])

            for ci in commission_idx:
                if ci in used_com:
                    continue
                c_amt = float(df.at[ci, "amount"])
                expected_initial = -(p_amt + c_amt)  # negative initial withdrawal
                di = find_debit(expected_initial)
                if di is None:
                    continue

                used_pr.add(pi)
                used_com.add(ci)
                used_deb.add(di)

                df.at[pi, "is_cancelled_forteforex_bundle"] = 1
                df.at[ci, "is_cancelled_forteforex_bundle"] = 1
                df.at[di, "is_cancelled_forteforex_bundle"] = 1
                break

    # 3) Write XLSX (transactions + summary)
    import pandas as pd  # type: ignore

    df_out = df.drop(columns=["date_only", "_text"])
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df_out.to_excel(xw, index=False, sheet_name="transactions")

    # 4) Add Excel formulas + summary sheet
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.styles import Font  # type: ignore
    from openpyxl.utils import get_column_letter

    wb = load_workbook(out_path)
    ws_tx = wb["transactions"]

    # Column lookup by header value
    header = {}
    for col in range(1, ws_tx.max_column + 1):
        v = ws_tx.cell(row=1, column=col).value
        if v:
            header[str(v)] = col

    def col_letter(name: str) -> str:
        return get_column_letter(header[name])

    col_amount = col_letter("amount")
    col_currency = col_letter("currency")
    col_cancel = col_letter("is_cancelled_forteforex_bundle")

    last_row = ws_tx.max_row
    B_rng = f"transactions!${col_amount}$2:${col_amount}${last_row}"
    C_rng = f"transactions!${col_currency}$2:${col_currency}${last_row}"
    I_rng = f"transactions!${col_cancel}$2:${col_cancel}${last_row}"

    ws = wb.create_sheet("summary") if "summary" not in wb.sheetnames else wb["summary"]

    # Overwrite key cells (keeps operation deterministic even if sheet existed)
    ws["A1"] = "KZT totals and cancellations adjustment"
    ws["A1"].font = Font(bold=True)
    ws.merge_cells("A1:B1")

    ws["A2"] = "Total incoming KZT (all)"
    ws["A3"] = "Total withdrawals KZT (all)"
    ws["A4"] = "Net KZT (all)"
    for r in (2, 3, 4):
        ws[f"A{r}"].font = Font(bold=True)

    ws["B2"] = f"=SUMIFS({B_rng},{C_rng},\"KZT\",{B_rng},\">0\")"
    ws["B3"] = f"=-SUMIFS({B_rng},{C_rng},\"KZT\",{B_rng},\"<0\")"
    ws["B4"] = "=B2-B3"

    ws["A6"] = "Net effect of cancelled ForteForex bundles (by flag)"
    ws["A7"] = "Net KZT excluding cancelled bundles"
    ws["A8"] = "Clean net check"
    for r in (6, 7, 8):
        ws[f"A{r}"].font = Font(bold=True)

    ws["B6"] = f"=SUMIFS({B_rng},{C_rng},\"KZT\",{I_rng},1)"
    ws["B7"] = "=B4-B6"
    ws["B8"] = "=B7"

    ws["A10"] = "Income/withdrawals excluding cancelled bundles"
    ws["A10"].font = Font(bold=True)
    ws.merge_cells("A10:B10")

    ws["A11"] = "Income (KZT) excl cancelled"
    ws["A12"] = "Withdrawals (KZT) excl cancelled"
    ws["A13"] = "Net (excl cancelled)"
    for r in (11, 12, 13):
        ws[f"A{r}"].font = Font(bold=True)

    ws["B11"] = f"=SUMIFS({B_rng},{C_rng},\"KZT\",{B_rng},\">0\",{I_rng},0)"
    ws["B12"] = f"=-SUMIFS({B_rng},{C_rng},\"KZT\",{B_rng},\"<0\",{I_rng},0)"
    ws["B13"] = "=B11-B12"

    ws["A15"] = "Initial balance at period start (KZT)"
    ws["A16"] = "Expected final balance (KZT)"
    ws["A17"] = "Final balance from statement (KZT)"
    for r in (15, 16, 17):
        ws[f"A{r}"].font = Font(bold=True)

    if args.initial_balance_kzt is not None:
        ws["B15"] = float(args.initial_balance_kzt)
    ws["B16"] = "=B15+B13"

    final_kzt = extract_final_balance_kzt(pdf_path)
    if final_kzt is not None:
        ws["B17"] = float(final_kzt)

    # Number formatting
    for cell in ("B2", "B3", "B4", "B6", "B7", "B8", "B11", "B12", "B13", "B15", "B16", "B17"):
        try:
            ws[cell].number_format = "#,##0.00"
        except Exception:
            pass

    ws.freeze_panes = "A3"
    wb.save(out_path)

    print(f"Wrote {len(df_out)} rows to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

